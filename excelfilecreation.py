# -*- coding: utf-8 -*-
"""
Created on Sun Jun 10 15:07:12 2018

@author: Balasubramaniam
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font,Fill,Border
from openpyxl.styles.colors import RED
import calendar
path="F:/AllState_Python2018/logs/reports/AllStateAnnualReport.xlsx"
wb=Workbook()
for month in calendar.month_name:
    print(month)
    if not(len(str(month))==0):
      wb.create_sheet(month+"_2018")
wb.save(path)
wb.close()

wbRef=load_workbook(path,read_only=False)
for sheet in wbRef.get_sheet_names():
    print(sheet)
import datetime    
currentMonth=datetime.date.today().strftime("%B")+"_2018"
sheetRef=wbRef.get_sheet_by_name(currentMonth)
import random


for row in range(1,100):
  for col in range(1,10):
      sheetRef.cell(column=col,row=row,value="%d" %(random.randint(100,10000)))
wbRef.save(path)
for row in range(1,100):
  for col in range(1,10):
      cell=sheetRef.cell(row=row,column=col)
      if(int(cell.value) > 5000):       
           cell.font = Font(size=18,color=RED)
wbRef.save(path)
wbRef.close()


    
      



